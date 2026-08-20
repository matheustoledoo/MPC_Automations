#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automação do portal da Receita Federal por aprendizado de mouse e teclado.

FASE 1 — PRIMEIRA EMPRESA, TOTALMENTE MANUAL:
1. O usuário faz todo o fluxo manualmente, sem qualquer gravação.
2. O programa orienta o usuário a preencher CNPJ e Procurador.
3. Antes de Representar, o programa impõe uma espera de 30 segundos.
4. O usuário clica em Representar, fecha o menu lateral e baixa o PDF.
5. O programa apenas detecta, move e renomeia o arquivo baixado.

FASE 2 — SEGUNDA EMPRESA, APRENDIZADO:
1. O usuário executa novamente o fluxo, agora com o programa registrando:
   perfil, campo CNPJ, ritmo da digitação, campo de perfil, Procurador,
   Representar e fechamento do menu lateral.
2. Antes do clique em Representar, o programa espera 30 segundos.
3. Para o download, o usuário marca DEZ posições possíveis do botão
   "Baixar Relatório" colocando o ponteiro em cada posição e pressionando F8.
   Nenhum clique é feito durante essa marcação.

FASE 3 — TERCEIRA EMPRESA EM DIANTE, AUTOMÁTICO:
1. Volta ao topo da página.
2. Clica no nome/perfil da empresa atualmente representada.
3. Clica no CNPJ, limpa e digita o novo documento.
4. Seleciona Procurador.
5. Espera obrigatoriamente 30 segundos.
6. Clica em Representar e fecha o menu lateral.
7. Força a página para o topo máximo e testa as dez posições do botão.
8. Antes de cada tentativa, força novamente a página para o topo.
9. Para imediatamente quando detectar que um PDF começou a baixar.
10. Copia o PDF renomeado para a pasta Saídas.
11. A pasta Saídas recebe SOMENTE PDFs; logs e CSV ficam no AppData local.

Esta versão:
- Não usa Selenium.
- Não usa ChromeDriver.
- Não usa depuração remota.
- Não usa PyAutoGUI, PyScreeze ou Pillow.
- Usa pynput para mouse e teclado.
- Ignora a META PLANO no processamento; ela é usada somente no login.
"""

from __future__ import annotations

import csv
import ctypes
import importlib.util
import json
import logging
import os
import random
import re
import shutil
import subprocess
import sys
import time
import traceback
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Sequence


# =============================================================================
# WINDOWS / DPI
# =============================================================================

def enable_windows_dpi_awareness() -> None:
    """Mantém as coordenadas corretas com escala do Windows em 125%, 150% etc."""
    if os.name != "nt":
        return

    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


enable_windows_dpi_awareness()


# =============================================================================
# CONFIGURAÇÕES PRINCIPAIS
# =============================================================================

DEFAULT_EXCEL_PATH = r"C:\Users\z0057vzy\OneDrive - Siemens Healthineers\Desktop\Teste\Planilha_Clientes-1784744405619.xlsx"
EXCEL_PATH = Path(os.environ.get("ECAC_EXCEL_PATH", DEFAULT_EXCEL_PATH))
EXCEL_FILENAME = EXCEL_PATH.name
SHEET_NAME = "📋 Cadastro de Clientes"

FIRST_DATA_ROW = 4
DOCUMENT_COLUMN = 4   # Coluna D
STATUS_COLUMN = 7     # Coluna G
NAME_COLUMN = 2       # Coluna B
CODE_COLUMN = 1       # Coluna A
ACTIVE_STATUS = "Ativo"

DEFAULT_OUTPUT_DIR = r"C:\Users\z0057vzy\OneDrive - Siemens Healthineers\Desktop\Saidas"
OUTPUT_DIR = Path(os.environ.get("ECAC_OUTPUT_DIR", DEFAULT_OUTPUT_DIR))

LOGIN_START_URL = "https://servicos.receitafederal.gov.br/servico/"

PORTAL_URL = (
    "https://servicos.receitafederal.gov.br/servico/"
    "pendencias/#/analise-pendencias"
)

META_PLANO_CNPJ = "22588294000157"
SKIP_META_PLANO_PROCESSING = True

AUTO_INSTALL_DEPENDENCIES = True

# Aguarde o carregamento depois de cada ação.
WAIT_AFTER_OPEN_PANEL_SECONDS = (1.0, 1.6)
WAIT_AFTER_CNPJ_SECONDS = (1.0, 1.7)
WAIT_AFTER_PROFILE_OPEN_SECONDS = (0.5, 0.9)
WAIT_AFTER_PROCURADOR_SECONDS = (0.7, 1.1)

# Espera obrigatória para evitar a mensagem do portal pedindo para aguardar.
WAIT_BEFORE_REPRESENT_SECONDS = 30

WAIT_AFTER_REPRESENT_SECONDS = (45.0, 45.0)
WAIT_AFTER_CLOSE_PANEL_SECONDS = (0.8, 1.4)

DELAY_BETWEEN_COMPANIES_SECONDS = (4.0, 7.0)
LONG_PAUSE_EVERY = 25
LONG_PAUSE_SECONDS = 35

MANUAL_ACTION_TIMEOUT_SECONDS = 300
DOWNLOAD_START_TIMEOUT_SECONDS = 35
DOWNLOAD_TIMEOUT_SECONDS = 240

# O botão "Baixar Relatório" pode aparecer em dez alturas diferentes.
DOWNLOAD_POSITION_COUNT = 16
DOWNLOAD_POSITION_TEST_TIMEOUT_SECONDS = 5
WAIT_AFTER_FORCE_TOP_SECONDS = (0.6, 1.0)

# Uma nova tentativa somente nas empresas posteriores ao aprendizado.
MAX_RETRIES_AFTER_LEARNING = 2

# Movimento físico do ponteiro.
MOUSE_MOVE_DURATION_SECONDS = (0.35, 0.65)
MOUSE_CLICK_HOLD_SECONDS = (0.055, 0.125)

# Usa o Chrome regular e o perfil normal do usuário.
# Para abrir o navegador manualmente, altere para False.
OPEN_REGULAR_CHROME_AUTOMATICALLY = True

# Opcional: informe uma pasta específica de downloads.
# Deixe None para detectar pelo perfil normal do Chrome/Windows.
CUSTOM_BROWSER_DOWNLOAD_DIR: Optional[Path] = None

LOCAL_APP_DATA = Path(os.environ.get("LOCALAPPDATA", str(Path.home())))

# Todos os arquivos técnicos da automação ficam fora de OUTPUT_DIR.
# A pasta "Saidas" fica reservada EXCLUSIVAMENTE aos PDFs finais.
RUNTIME_DIR = (
    LOCAL_APP_DATA
    / "ReceitaFederalECAC"
    / "runtime"
)


# =============================================================================
# DEPENDÊNCIAS
# =============================================================================

REQUIRED_PACKAGES = {
    "openpyxl": "openpyxl>=3.1",
    "pypdf": "pypdf>=5.0",
    "pynput": "pynput>=1.7.7",
}


def ensure_dependencies() -> None:
    if os.name != "nt":
        raise RuntimeError("Esta automação foi criada para Windows 11.")

    missing = [
        requirement
        for module_name, requirement in REQUIRED_PACKAGES.items()
        if importlib.util.find_spec(module_name) is None
    ]

    if not missing:
        return

    if not AUTO_INSTALL_DEPENDENCIES:
        raise RuntimeError(
            "Bibliotecas ausentes: " + ", ".join(missing)
        )

    print("Instalando bibliotecas necessárias: " + ", ".join(missing))
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "--upgrade", *missing]
    )


ensure_dependencies()

from openpyxl import load_workbook  # noqa: E402
from pypdf import PdfReader  # noqa: E402
from pynput import keyboard, mouse  # noqa: E402


# =============================================================================
# MODELOS E ESTADO
# =============================================================================

@dataclass(frozen=True)
class Client:
    excel_row: int
    code: str
    spreadsheet_name: str
    cnpj: str


class PortalError(RuntimeError):
    """Erro recuperável de uma empresa."""


LOGGER = logging.getLogger("receita_ecac_loop")
RUN_STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

LOG_FILE: Optional[Path] = None
RESULT_CSV_FILE: Optional[Path] = None
# Pasta permanente onde o Chrome mantém os arquivos originais.
TEMP_DOWNLOAD_DIR: Optional[Path] = None

# As ações normais guardam somente a posição absoluta do clique.
# Rolagens acidentais não são mais gravadas nem repetidas.
MOUSE_ACTIONS: dict[str, dict[str, Any]] = {}

# Cinco coordenadas possíveis do botão "Baixar Relatório".
# Elas são marcadas na segunda empresa usando a tecla F8, sem clicar.
DOWNLOAD_POSITIONS: list[tuple[int, int]] = []

ACTION_LABELS = {
    "open_representation_panel": (
        "nome/perfil que abre o painel Representar"
    ),
    "cnpj_field": "campo de CNPJ",
    "profile_dropdown": "campo de perfil",
    "procurador_option": "opção Procurador",
    "represent_button": "botão Representar",
    "close_representation_panel": "X para fechar o menu lateral",
}

# Somente estas ações fazem parte do fluxo de representação.
# Outras automações da Central podem acrescentar rótulos em ACTION_LABELS
# para reaproveitar o mesmo aprendizado sem alterar o que é exigido aqui.
REQUIRED_ACTION_KEYS = frozenset(ACTION_LABELS)

# Ritmo registrado na digitação manual do primeiro CNPJ.
TYPING_PATTERN: list[dict[str, float]] = []


# =============================================================================
# UTILITÁRIOS
# =============================================================================

def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(
        char
        for char in text
        if not unicodedata.combining(char)
    )
    return re.sub(r"\s+", " ", text).strip().casefold()


def only_digits(value: object) -> str:
    return re.sub(r"\D", "", "" if value is None else str(value))


def format_cnpj(cnpj: str) -> str:
    digits = only_digits(cnpj).zfill(14)
    if len(digits) != 14:
        return cnpj

    return (
        f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/"
        f"{digits[8:12]}-{digits[12:]}"
    )


def is_valid_cnpj(cnpj: str) -> bool:
    digits = only_digits(cnpj)

    if len(digits) != 14 or digits == digits[0] * 14:
        return False

    def calculate_digit(
        base: str,
        weights: Sequence[int],
    ) -> str:
        total = sum(
            int(number) * weight
            for number, weight in zip(base, weights)
        )
        remainder = total % 11
        return "0" if remainder < 2 else str(11 - remainder)

    first = calculate_digit(
        digits[:12],
        [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2],
    )
    second = calculate_digit(
        digits[:12] + first,
        [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2],
    )

    return digits[-2:] == first + second


def sanitize_filename(
    text: str,
    max_length: int = 125,
) -> str:
    text = re.sub(r'[<>:"/\\|?*\x00-\x1F]', " ", text)
    text = re.sub(r"\s+", " ", text).strip(" .")

    if not text:
        text = "EMPRESA SEM NOME"

    return text[:max_length].rstrip(" .")


def random_sleep(interval: tuple[float, float]) -> None:
    time.sleep(random.uniform(*interval))


def resolve_excel_path() -> Path:
    candidates = [
        EXCEL_PATH,
        Path(__file__).resolve().with_name(EXCEL_FILENAME),
        Path.cwd() / EXCEL_FILENAME,
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    raise FileNotFoundError(
        "Planilha não encontrada. Caminhos verificados:\n- "
        + "\n- ".join(str(path) for path in candidates)
    )


def initialize_folders_and_logging() -> None:
    global LOG_FILE, RESULT_CSV_FILE, TEMP_DOWNLOAD_DIR

    # IMPORTANTE:
    # OUTPUT_DIR é uma pasta "limpa" de integração.
    # A automação NUNCA cria log, CSV, diagnóstico ou subpasta nela.
    # Somente os PDFs finais são gravados em OUTPUT_DIR.
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Arquivos técnicos ficam no AppData local, fora do OneDrive/Saidas.
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)

    # Observa a pasta real de downloads do Chrome normal.
    # O PDF original baixa na pasta Downloads do navegador e somente a cópia
    # renomeada é colocada em OUTPUT_DIR.
    TEMP_DOWNLOAD_DIR = resolve_browser_download_directory()
    TEMP_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    LOG_FILE = RUNTIME_DIR / f"execucao_receita_loop_{RUN_STAMP}.log"
    RESULT_CSV_FILE = (
        RUNTIME_DIR / f"resultado_receita_loop_{RUN_STAMP}.csv"
    )

    LOGGER.setLevel(logging.INFO)
    LOGGER.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(
        LOG_FILE,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)

    LOGGER.addHandler(file_handler)
    LOGGER.addHandler(console_handler)


def load_clients(excel_path: Path) -> list[Client]:
    workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=True,
    )

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise KeyError(
                f"A aba '{SHEET_NAME}' não foi encontrada. "
                f"Abas disponíveis: {', '.join(workbook.sheetnames)}"
            )

        sheet = workbook[SHEET_NAME]
        clients: list[Client] = []
        invalid_rows: list[int] = []
        skipped_meta_rows: list[int] = []

        for row_number in range(
            FIRST_DATA_ROW,
            sheet.max_row + 1,
        ):
            status = sheet.cell(
                row_number,
                STATUS_COLUMN,
            ).value

            if normalize_text(status) != normalize_text(ACTIVE_STATUS):
                continue

            cnpj = only_digits(
                sheet.cell(
                    row_number,
                    DOCUMENT_COLUMN,
                ).value
            )

            # Ignora CPFs e outros documentos.
            if len(cnpj) != 14:
                continue

            if (
                SKIP_META_PLANO_PROCESSING
                and cnpj == META_PLANO_CNPJ
            ):
                skipped_meta_rows.append(row_number)
                continue

            if not is_valid_cnpj(cnpj):
                invalid_rows.append(row_number)

            clients.append(
                Client(
                    excel_row=row_number,
                    code=str(
                        sheet.cell(
                            row_number,
                            CODE_COLUMN,
                        ).value
                        or ""
                    ).strip(),
                    spreadsheet_name=str(
                        sheet.cell(
                            row_number,
                            NAME_COLUMN,
                        ).value
                        or ""
                    ).strip(),
                    cnpj=cnpj,
                )
            )

        if skipped_meta_rows:
            LOGGER.info(
                "META PLANO ignorada no processamento "
                "(linha(s) %s).",
                ", ".join(map(str, skipped_meta_rows)),
            )

        if invalid_rows:
            LOGGER.warning(
                "CNPJs com dígito verificador inconsistente "
                "nas linhas %s. Eles continuarão na lista.",
                ", ".join(map(str, invalid_rows)),
            )

        clients.sort(key=lambda client: client.excel_row)
        return clients

    finally:
        workbook.close()


# =============================================================================
# CSV DE RESULTADO
# =============================================================================

def create_result_csv() -> tuple[Any, csv.DictWriter]:
    if RESULT_CSV_FILE is None:
        raise RuntimeError("CSV de resultado não configurado.")

    handle = RESULT_CSV_FILE.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    )

    fieldnames = [
        "linha_excel",
        "codigo",
        "nome_planilha",
        "cnpj",
        "status",
        "tentativa",
        "mensagem",
        "nome_extraido_relatorio",
        "arquivo_final",
        "inicio",
        "fim",
    ]

    writer = csv.DictWriter(
        handle,
        fieldnames=fieldnames,
        delimiter=";",
    )
    writer.writeheader()
    handle.flush()

    return handle, writer


def write_result(
    handle: Any,
    writer: csv.DictWriter,
    client: Client,
    status: str,
    attempt: int,
    message: str,
    started_at: datetime,
    report_name: str = "",
    final_path: str = "",
) -> None:
    writer.writerow(
        {
            "linha_excel": client.excel_row,
            "codigo": client.code,
            "nome_planilha": client.spreadsheet_name,
            "cnpj": format_cnpj(client.cnpj),
            "status": status,
            "tentativa": attempt,
            "mensagem": message,
            "nome_extraido_relatorio": report_name,
            "arquivo_final": final_path,
            "inicio": started_at.strftime("%Y-%m-%d %H:%M:%S"),
            "fim": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    handle.flush()


# =============================================================================
# CHROME NORMAL
# =============================================================================

def find_chrome_executable() -> Path:
    candidates = [
        Path(
            os.environ.get(
                "PROGRAMFILES",
                r"C:\Program Files",
            )
        )
        / "Google"
        / "Chrome"
        / "Application"
        / "chrome.exe",
        Path(
            os.environ.get(
                "PROGRAMFILES(X86)",
                r"C:\Program Files (x86)",
            )
        )
        / "Google"
        / "Chrome"
        / "Application"
        / "chrome.exe",
        Path(
            os.environ.get(
                "LOCALAPPDATA",
                str(Path.home()),
            )
        )
        / "Google"
        / "Chrome"
        / "Application"
        / "chrome.exe",
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    chrome_from_path = (
        shutil.which("chrome")
        or shutil.which("chrome.exe")
    )

    if chrome_from_path:
        return Path(chrome_from_path)

    raise RuntimeError("Google Chrome não encontrado.")


def resolve_browser_download_directory() -> Path:
    """Detecta a pasta usada pelo perfil normal do Chrome."""
    if CUSTOM_BROWSER_DOWNLOAD_DIR is not None:
        path = Path(
            os.path.expandvars(str(CUSTOM_BROWSER_DOWNLOAD_DIR))
        ).expanduser()
        LOGGER.info(
            "Pasta de downloads configurada manualmente: %s",
            path,
        )
        return path

    chrome_user_data = (
        LOCAL_APP_DATA / "Google" / "Chrome" / "User Data"
    )
    profile_name = "Default"

    local_state_path = chrome_user_data / "Local State"
    if local_state_path.exists():
        try:
            local_state = json.loads(
                local_state_path.read_text(
                    encoding="utf-8",
                    errors="replace",
                )
            )
            profile_name = (
                local_state.get("profile", {}).get("last_used")
                or profile_name
            )
        except Exception as exc:
            LOGGER.warning(
                "Não foi possível identificar o último perfil do Chrome: %s",
                exc,
            )

    preferences_path = (
        chrome_user_data / profile_name / "Preferences"
    )
    if preferences_path.exists():
        try:
            preferences = json.loads(
                preferences_path.read_text(
                    encoding="utf-8",
                    errors="replace",
                )
            )
            configured_directory = (
                preferences.get("download", {}).get(
                    "default_directory"
                )
            )

            if configured_directory:
                path = Path(
                    os.path.expandvars(str(configured_directory))
                ).expanduser()

                LOGGER.info(
                    "Pasta de downloads detectada no perfil normal "
                    "do Chrome (%s): %s",
                    profile_name,
                    path,
                )
                return path
        except Exception as exc:
            LOGGER.warning(
                "Não foi possível ler as preferências do Chrome: %s",
                exc,
            )

    try:
        import winreg

        registry_path = (
            r"Software\Microsoft\Windows\CurrentVersion\Explorer"
            r"\User Shell Folders"
        )
        downloads_guid = (
            "{374DE290-123F-4565-9164-39C4925E467B}"
        )

        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            registry_path,
        ) as key:
            value, _ = winreg.QueryValueEx(
                key,
                downloads_guid,
            )

        path = Path(
            os.path.expandvars(str(value))
        ).expanduser()

        LOGGER.info(
            "Pasta de downloads detectada pelo Windows: %s",
            path,
        )
        return path

    except Exception as exc:
        LOGGER.warning(
            "Não foi possível consultar a pasta Downloads do Windows: %s",
            exc,
        )

    fallback = Path.home() / "Downloads"
    LOGGER.warning(
        "Usando a pasta de downloads padrão: %s",
        fallback,
    )
    return fallback


def open_regular_chrome() -> None:
    """
    Abre o Chrome instalado com o perfil normal do usuário.

    Não cria perfil separado, não edita Preferences, não usa depuração
    remota e não usa Selenium ou ChromeDriver.
    """
    if not OPEN_REGULAR_CHROME_AUTOMATICALLY:
        LOGGER.info(
            "A abertura automática está desativada. "
            "Abra o Chrome normal manualmente."
        )
        return

    chrome = find_chrome_executable()

    LOGGER.info(
        "Abrindo o Chrome regular com o perfil normal do usuário."
    )

    try:
        subprocess.Popen(
            [str(chrome), LOGIN_START_URL],
            close_fds=True,
        )
    except Exception as exc:
        raise RuntimeError(
            f"Não foi possível abrir o Chrome regular: {exc}"
        ) from exc

    time.sleep(3)


# =============================================================================
# TECLADO FÍSICO
# =============================================================================

def press_and_release_key(
    controller: keyboard.Controller,
    key: keyboard.Key | keyboard.KeyCode | str,
    hold_seconds: float,
) -> None:
    controller.press(key)
    time.sleep(hold_seconds)
    controller.release(key)


def keyboard_ctrl_home() -> None:
    """
    Força a página para o topo máximo.

    O comando é enviado três vezes para neutralizar qualquer rolagem residual,
    restauração de posição feita pelo navegador ou clique que tenha alterado
    levemente o scroll da página.
    """
    controller = keyboard.Controller()

    for _ in range(3):
        ctrl_pressed = False
        home_pressed = False

        try:
            controller.press(keyboard.Key.ctrl)
            ctrl_pressed = True
            time.sleep(0.05)

            controller.press(keyboard.Key.home)
            home_pressed = True
            time.sleep(0.08)

            controller.release(keyboard.Key.home)
            home_pressed = False
            time.sleep(0.03)

            controller.release(keyboard.Key.ctrl)
            ctrl_pressed = False
        finally:
            # Evita deixar Ctrl ou Home pressionados em caso de erro.
            if home_pressed:
                try:
                    controller.release(keyboard.Key.home)
                except Exception:
                    pass

            if ctrl_pressed:
                try:
                    controller.release(keyboard.Key.ctrl)
                except Exception:
                    pass

        time.sleep(0.12)

    random_sleep(WAIT_AFTER_FORCE_TOP_SECONDS)


def navigate_to_portal_with_keyboard() -> None:
    """Reabre a página inicial do serviço usando Ctrl+L."""
    controller = keyboard.Controller()

    controller.press(keyboard.Key.ctrl)
    time.sleep(0.05)
    controller.press("l")
    time.sleep(0.06)
    controller.release("l")
    controller.release(keyboard.Key.ctrl)

    time.sleep(0.12)

    for char in PORTAL_URL:
        press_and_release_key(
            controller,
            char,
            0.008,
        )

    press_and_release_key(
        controller,
        keyboard.Key.enter,
        0.05,
    )

    time.sleep(10)


def wait_before_represent() -> None:
    """Espera exatamente 30 segundos antes de permitir o clique em Representar."""
    LOGGER.info(
        "Aguardando %d segundos antes de clicar em 'Representar'.",
        WAIT_BEFORE_REPRESENT_SECONDS,
    )

    remaining = WAIT_BEFORE_REPRESENT_SECONDS

    while remaining > 0:
        LOGGER.info(
            "Representar será liberado em %d segundo(s).",
            remaining,
        )
        step = min(10, remaining)
        time.sleep(step)
        remaining -= step

    LOGGER.info(
        "Espera concluída. O clique em 'Representar' está liberado."
    )


def clear_cnpj_field() -> None:
    """Executa Ctrl+A e Backspace fisicamente."""
    controller = keyboard.Controller()

    controller.press(keyboard.Key.ctrl)
    time.sleep(random.uniform(0.045, 0.085))

    controller.press("a")
    time.sleep(random.uniform(0.045, 0.095))
    controller.release("a")

    time.sleep(random.uniform(0.025, 0.060))
    controller.release(keyboard.Key.ctrl)

    time.sleep(random.uniform(0.10, 0.18))

    press_and_release_key(
        controller,
        keyboard.Key.backspace,
        random.uniform(0.045, 0.095),
    )

    time.sleep(random.uniform(0.14, 0.24))


# =============================================================================
# MOUSE: APRENDER E REPETIR
# =============================================================================

def capture_mouse_action(
    action_key: str,
    instruction: str,
) -> None:
    """
    Registra somente a posição do clique esquerdo.

    A versão anterior também gravava movimentos da roda. Uma rolagem
    involuntária durante o aprendizado era reproduzida no começo das empresas,
    causando o scroll inesperado relatado.
    """
    label = ACTION_LABELS[action_key]

    print()
    print("=" * 94)
    print(f"APRENDENDO: {label.upper()}")
    print("=" * 94)
    print(instruction)
    print("Não use a roda do mouse durante esta etapa.")
    print(
        "Faça somente o clique solicitado. "
        "Esse clique já executará a ação e salvará a coordenada."
    )
    print("=" * 94)

    captured: dict[str, Any] = {
        "position": None,
    }

    def on_click(
        x: int,
        y: int,
        button: mouse.Button,
        pressed: bool,
    ) -> Optional[bool]:
        if pressed and button == mouse.Button.left:
            captured["position"] = (
                int(x),
                int(y),
            )
            return False

        return None

    LOGGER.warning(
        "Aguardando clique manual para '%s' por até %d segundos.",
        label,
        MANUAL_ACTION_TIMEOUT_SECONDS,
    )

    listener = mouse.Listener(on_click=on_click)
    listener.start()
    listener.join(MANUAL_ACTION_TIMEOUT_SECONDS)

    if listener.is_alive():
        listener.stop()
        raise PortalError(
            f"Tempo esgotado aguardando: {label}."
        )

    position = captured["position"]

    if position is None:
        raise PortalError(
            f"Nenhum clique foi registrado para: {label}."
        )

    MOUSE_ACTIONS[action_key] = captured

    LOGGER.info(
        "Posição aprendida para '%s': X=%d, Y=%d.",
        label,
        position[0],
        position[1],
    )


def move_mouse_smoothly(
    controller: mouse.Controller,
    target_x: int,
    target_y: int,
) -> None:
    start_x, start_y = controller.position
    duration = random.uniform(*MOUSE_MOVE_DURATION_SECONDS)

    steps = max(15, int(duration / 0.012))

    for step in range(1, steps + 1):
        progress = step / steps

        # Suavização simples para evitar salto instantâneo.
        eased = progress * progress * (3 - 2 * progress)

        current_x = round(
            start_x + (target_x - start_x) * eased
        )
        current_y = round(
            start_y + (target_y - start_y) * eased
        )

        controller.position = (
            current_x,
            current_y,
        )

        time.sleep(duration / steps)


def replay_mouse_action(action_key: str) -> None:
    action = MOUSE_ACTIONS[action_key]
    target_x, target_y = action["position"]

    controller = mouse.Controller()

    move_mouse_smoothly(
        controller,
        target_x,
        target_y,
    )

    controller.press(mouse.Button.left)
    time.sleep(random.uniform(*MOUSE_CLICK_HOLD_SECONDS))
    controller.release(mouse.Button.left)


def perform_mouse_action(
    action_key: str,
    instruction: str,
) -> bool:
    """
    Retorna True se a ação foi aprendida agora.
    Retorna False se foi repetida automaticamente.
    """
    if action_key not in MOUSE_ACTIONS:
        capture_mouse_action(
            action_key,
            instruction,
        )
        return True

    LOGGER.info(
        "Repetindo fisicamente: %s.",
        ACTION_LABELS[action_key],
    )

    replay_mouse_action(action_key)
    return False


def capture_download_position(position_number: int) -> None:
    """
    Marca uma posição possível sem clicar no portal.

    O usuário posiciona o ponteiro e pressiona F8. A tela permanece no
    topo e nenhum comando de rolagem para baixo é executado.
    """
    print()
    print("=" * 94)
    print(
        f"MARCAR POSIÇÃO {position_number}/{DOWNLOAD_POSITION_COUNT} "
        "DO BOTÃO BAIXAR RELATÓRIO"
    )
    print("=" * 94)
    print(
        "Posicione o ponteiro sobre uma das posições possíveis do botão "
        "'Baixar Relatório'."
    )
    print("NÃO clique. Quando o ponteiro estiver no lugar, pressione F8.")
    print(
        "Marque dez posições diferentes, especialmente variando a altura "
        "vertical em que o botão pode aparecer."
    )
    print("=" * 94)

    captured_position: dict[str, Optional[tuple[int, int]]] = {
        "value": None,
    }

    def on_press(
        key: keyboard.Key | keyboard.KeyCode,
    ) -> Optional[bool]:
        if key == keyboard.Key.f8:
            x, y = mouse.Controller().position
            captured_position["value"] = (int(x), int(y))
            return False

        if key == keyboard.Key.esc:
            return False

        return None

    LOGGER.warning(
        "Aguardando F8 para marcar a posição %d/%d por até %d segundos.",
        position_number,
        DOWNLOAD_POSITION_COUNT,
        MANUAL_ACTION_TIMEOUT_SECONDS,
    )

    listener = keyboard.Listener(on_press=on_press)
    listener.start()
    listener.join(MANUAL_ACTION_TIMEOUT_SECONDS)

    if listener.is_alive():
        listener.stop()
        raise PortalError(
            "Tempo esgotado durante a marcação das posições de download."
        )

    position = captured_position["value"]

    if position is None:
        raise PortalError(
            "A marcação foi cancelada ou F8 não foi pressionado."
        )

    DOWNLOAD_POSITIONS.append(position)

    LOGGER.info(
        "Posição de download %d/%d marcada: X=%d, Y=%d.",
        position_number,
        DOWNLOAD_POSITION_COUNT,
        position[0],
        position[1],
    )


def capture_all_download_positions() -> None:
    """Marca exatamente dez posições possíveis do botão de download."""
    if len(DOWNLOAD_POSITIONS) == DOWNLOAD_POSITION_COUNT:
        return

    DOWNLOAD_POSITIONS.clear()

    # Todas as posições devem ser marcadas com a página no topo máximo.
    keyboard_ctrl_home()

    print()
    print("=" * 94)
    print("APRENDIZADO DAS DEZ POSIÇÕES DE DOWNLOAD")
    print("=" * 94)
    print(
        "A página foi forçada para o topo máximo. Agora marque cinco "
        "pontos possíveis para o botão 'Baixar Relatório'."
    )
    print(
        "Em cada ponto, mantenha a página no topo, coloque o ponteiro "
        "sobre a posição e pressione F8."
    )
    print("Não clique durante a marcação.")
    print("=" * 94)

    for position_number in range(1, DOWNLOAD_POSITION_COUNT + 1):
        capture_download_position(position_number)

    LOGGER.info(
        "As %d posições possíveis do botão de download foram registradas.",
        DOWNLOAD_POSITION_COUNT,
    )


def click_download_positions_until_started(
    download_dir: Path,
    files_before: set[Path],
    started_timestamp: float,
) -> None:
    """
    Testa as dez posições na ordem em que foram marcadas, sempre
    forçando a página para o topo antes de cada tentativa.

    Após cada clique, aguarda alguns segundos. Assim que um arquivo novo
    aparecer, interrompe o teste e não clica nas posições restantes.
    """
    if len(DOWNLOAD_POSITIONS) != DOWNLOAD_POSITION_COUNT:
        raise PortalError(
            "As dez posições do botão 'Baixar Relatório' "
            "ainda não foram aprendidas."
        )

    keyboard_ctrl_home()
    controller = mouse.Controller()

    for index, (target_x, target_y) in enumerate(
        DOWNLOAD_POSITIONS,
        start=1,
    ):
        # Garante que cada posição seja testada a partir do mesmo topo.
        keyboard_ctrl_home()

        LOGGER.info(
            "Testando posição %d/%d de 'Baixar Relatório' com a página no topo: "
            "X=%d, Y=%d.",
            index,
            DOWNLOAD_POSITION_COUNT,
            target_x,
            target_y,
        )

        move_mouse_smoothly(
            controller,
            target_x,
            target_y,
        )

        controller.press(mouse.Button.left)
        time.sleep(random.uniform(*MOUSE_CLICK_HOLD_SECONDS))
        controller.release(mouse.Button.left)

        if wait_for_download_start(
            download_dir,
            files_before,
            started_timestamp,
            timeout=DOWNLOAD_POSITION_TEST_TIMEOUT_SECONDS,
        ):
            LOGGER.info(
                "Download detectado após clicar na posição %d/%d.",
                index,
                DOWNLOAD_POSITION_COUNT,
            )
            return

        LOGGER.warning(
            "Nenhum download detectado na posição %d/%d. "
            "Tentando a próxima.",
            index,
            DOWNLOAD_POSITION_COUNT,
        )

        # Um clique fora do botão pode alterar levemente o foco/rolagem.
        # Retorna imediatamente ao topo máximo antes da próxima coordenada.
        keyboard_ctrl_home()

    raise PortalError(
        "As dez posições de 'Baixar Relatório' foram testadas, "
        "mas nenhum download começou."
    )


# =============================================================================
# APRENDIZADO DA DIGITAÇÃO
# =============================================================================

def capture_first_cnpj_and_typing_pattern(
    expected_cnpj: str,
) -> None:
    """
    O usuário digita o primeiro CNPJ manualmente.

    São registrados:
    - intervalo entre as teclas;
    - duração de cada tecla pressionada.

    Ctrl+A e Backspace podem ser usados antes da digitação.
    Eles não entram no padrão dos 14 números.
    """
    global TYPING_PATTERN

    expected = only_digits(expected_cnpj)

    if len(expected) != 14:
        raise PortalError(
            "O CNPJ esperado não possui 14 dígitos."
        )

    print()
    print("=" * 94)
    print("DIGITAÇÃO MANUAL DO PRIMEIRO CNPJ")
    print("=" * 94)
    print(
        f"Digite manualmente, número por número: "
        f"{format_cnpj(expected)}"
    )
    print(
        "Caso exista algum conteúdo no campo, você pode usar "
        "Ctrl+A e Backspace antes de digitar."
    )
    print(
        "Não copie e não cole. O programa registrará o ritmo "
        "das 14 teclas numéricas."
    )
    print(
        "Ao terminar corretamente, o programa seguirá "
        "automaticamente para o próximo clique."
    )
    print("=" * 94)

    events: list[dict[str, Any]] = []
    pressed_indexes: dict[str, list[int]] = {}
    last_press_time: Optional[float] = None

    def digit_from_key(
        key: keyboard.Key | keyboard.KeyCode,
    ) -> Optional[str]:
        try:
            char = key.char
        except AttributeError:
            return None

        if char and char.isdigit():
            return char

        return None

    def on_press(
        key: keyboard.Key | keyboard.KeyCode,
    ) -> Optional[bool]:
        nonlocal last_press_time

        if key == keyboard.Key.backspace:
            if events:
                events.pop()
                last_press_time = (
                    float(events[-1]["press_time"])
                    if events
                    else None
                )
            return None

        digit = digit_from_key(key)

        if digit is None:
            return None

        # Ignora repetição automática da mesma tecla mantida pressionada.
        open_indexes = pressed_indexes.setdefault(digit, [])

        if open_indexes:
            return None

        now = time.perf_counter()
        delay_before = (
            0.0
            if last_press_time is None
            else now - last_press_time
        )

        events.append(
            {
                "digit": digit,
                "press_time": now,
                "delay_before": delay_before,
                "hold": 0.055,
            }
        )

        open_indexes.append(len(events) - 1)
        last_press_time = now

        return None

    def on_release(
        key: keyboard.Key | keyboard.KeyCode,
    ) -> Optional[bool]:
        digit = digit_from_key(key)

        if digit is not None:
            indexes = pressed_indexes.get(digit, [])

            if indexes:
                index = indexes.pop()

                if 0 <= index < len(events):
                    events[index]["hold"] = max(
                        0.020,
                        time.perf_counter()
                        - float(events[index]["press_time"]),
                    )

        typed = "".join(
            str(event["digit"])
            for event in events
        )

        if len(typed) == 14:
            if typed == expected:
                return False

            LOGGER.warning(
                "Foram digitados 14 números, mas eles não "
                "correspondem ao CNPJ esperado. "
                "Corrija usando Backspace."
            )

        return None

    listener = keyboard.Listener(
        on_press=on_press,
        on_release=on_release,
    )
    listener.start()
    listener.join(MANUAL_ACTION_TIMEOUT_SECONDS)

    if listener.is_alive():
        listener.stop()
        raise PortalError(
            "Tempo esgotado aguardando a digitação "
            "manual do primeiro CNPJ."
        )

    typed = "".join(
        str(event["digit"])
        for event in events
    )

    if typed != expected:
        raise PortalError(
            f"O CNPJ registrado foi {typed!r}, "
            f"mas o esperado era {expected!r}."
        )

    TYPING_PATTERN = [
        {
            "delay_before": min(
                max(
                    float(event["delay_before"]),
                    0.035,
                ),
                0.600,
            ),
            "hold": min(
                max(
                    float(event["hold"]),
                    0.025,
                ),
                0.220,
            ),
        }
        for event in events
    ]

    LOGGER.info(
        "Ritmo manual aprendido com 14 teclas."
    )


def type_cnpj_with_learned_pattern(cnpj: str) -> None:
    digits = only_digits(cnpj)

    if len(digits) != 14:
        raise PortalError(
            f"CNPJ inválido para digitação: {cnpj!r}"
        )

    if len(TYPING_PATTERN) != 14:
        raise PortalError(
            "O ritmo de digitação ainda não foi aprendido."
        )

    clear_cnpj_field()
    controller = keyboard.Controller()

    for index, digit in enumerate(digits):
        pattern = TYPING_PATTERN[index]

        if index > 0:
            delay = (
                float(pattern["delay_before"])
                * random.uniform(0.90, 1.10)
            )
            time.sleep(max(0.035, delay))

        hold = (
            float(pattern["hold"])
            * random.uniform(0.92, 1.08)
        )

        press_and_release_key(
            controller,
            digit,
            min(max(hold, 0.025), 0.230),
        )

    LOGGER.info(
        "CNPJ digitado com o ritmo aprendido: %s.",
        format_cnpj(cnpj),
    )


# =============================================================================
# DOWNLOAD
# =============================================================================

def list_download_files(
    download_dir: Path,
) -> set[Path]:
    try:
        return {
            path.resolve()
            for path in download_dir.iterdir()
            if path.is_file()
        }
    except OSError:
        return set()


def wait_for_download_start(
    download_dir: Path,
    files_before: set[Path],
    started_timestamp: float,
    timeout: int = DOWNLOAD_START_TIMEOUT_SECONDS,
) -> bool:
    deadline = time.time() + timeout

    while time.time() < deadline:
        current_files = list_download_files(download_dir)
        new_files = current_files - files_before

        for path in new_files:
            try:
                if (
                    path.stat().st_mtime
                    >= started_timestamp - 2
                ):
                    return True
            except OSError:
                continue

        time.sleep(0.5)

    return False


def wait_for_new_pdf(
    download_dir: Path,
    files_before: set[Path],
    started_timestamp: float,
    timeout: int = DOWNLOAD_TIMEOUT_SECONDS,
) -> Path:
    deadline = time.time() + timeout

    last_size: dict[Path, int] = {}
    stable_count: dict[Path, int] = {}

    while time.time() < deadline:
        partial_files = (
            list(download_dir.glob("*.crdownload"))
            + list(download_dir.glob("*.tmp"))
        )

        candidates: list[Path] = []

        for path in download_dir.glob("*.pdf"):
            try:
                if (
                    path.resolve() not in files_before
                    and path.stat().st_mtime
                    >= started_timestamp - 2
                ):
                    candidates.append(path)
            except OSError:
                continue

        candidates.sort(
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )

        for candidate in candidates:
            try:
                size = candidate.stat().st_size
            except OSError:
                continue

            if (
                size > 0
                and last_size.get(candidate) == size
            ):
                stable_count[candidate] = (
                    stable_count.get(candidate, 0) + 1
                )
            else:
                stable_count[candidate] = 0
                last_size[candidate] = size

            if (
                stable_count[candidate] >= 2
                and not partial_files
            ):
                return candidate

        time.sleep(1)

    raise PortalError(
        f"O PDF não terminou de baixar "
        f"em {timeout} segundos."
    )


# =============================================================================
# PDF E RENOMEAÇÃO
# =============================================================================

def clean_company_name_candidate(
    value: str,
) -> Optional[str]:
    value = re.sub(r"\s+", " ", value).strip(" :-")
    normalized = normalize_text(value)

    if len(value) < 5 or len(value) > 180:
        return None

    if sum(char.isalpha() for char in value) < 4:
        return None

    forbidden_terms = (
        "ministerio da fazenda",
        "secretaria especial da receita federal",
        "receita federal do brasil",
        "relatorio de situacao fiscal",
        "situacao fiscal",
        "detalhamento da analise",
        "data de emissao",
        "pagina",
        "cnpj",
        "cpf",
        "procuracao",
        "diagnostico fiscal",
        "pendencia",
        "servicos da receita federal",
    )

    if any(
        term in normalized
        for term in forbidden_terms
    ):
        return None

    if re.fullmatch(r"[\d\W_]+", value):
        return None

    return value


def extract_report_company_name(
    pdf_path: Path,
    cnpj: str,
) -> Optional[str]:
    try:
        reader = PdfReader(str(pdf_path))
        text_parts: list[str] = []

        for page in reader.pages[:4]:
            extracted = page.extract_text() or ""

            if extracted:
                text_parts.append(extracted)

        full_text = "\n".join(text_parts)

    except Exception as exc:
        LOGGER.warning(
            "Não foi possível ler o PDF %s: %s",
            pdf_path.name,
            exc,
        )
        return None

    if not full_text.strip():
        return None

    explicit_patterns = [
        (
            r"(?im)^\s*(?:nome\s+empresarial|"
            r"raz[aã]o\s+social|nome\s+do\s+contribuinte)"
            r"\s*[:\-]?\s*(.+?)\s*$"
        ),
        (
            r"(?im)^\s*(?:contribuinte)"
            r"\s*[:\-]\s*(.+?)\s*$"
        ),
    ]

    for pattern in explicit_patterns:
        match = re.search(pattern, full_text)

        if match:
            candidate = clean_company_name_candidate(
                match.group(1)
            )

            if candidate:
                return candidate

    lines = [
        re.sub(r"\s+", " ", line).strip()
        for line in full_text.splitlines()
    ]
    lines = [line for line in lines if line]

    cnpj_digits = only_digits(cnpj)

    indexes = [
        index
        for index, line in enumerate(lines)
        if (
            cnpj_digits
            and cnpj_digits in only_digits(line)
        )
    ]

    for index in indexes:
        for offset in (-1, -2, -3, 1, 2, -4, 3):
            target = index + offset

            if 0 <= target < len(lines):
                candidate = clean_company_name_candidate(
                    lines[target]
                )

                if candidate:
                    return candidate

    for line in lines:
        candidate = clean_company_name_candidate(line)

        if candidate:
            return candidate

    return None


def report_date_from_download_name(
    pdf_path: Path,
) -> str:
    match = re.search(
        r"-(\d{8})(?:\D|$)",
        pdf_path.stem,
    )

    if match:
        try:
            parsed = datetime.strptime(
                match.group(1),
                "%Y%m%d",
            )
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            pass

    return datetime.now().strftime("%Y-%m-%d")


def pdf_belongs_to_client(
    pdf_path: Path,
    client: Client,
) -> bool:
    """
    Evita dar o nome de uma empresa ao PDF de outra.

    Primeiro verifica o CNPJ no nome original. Se necessário, procura o CNPJ
    nas quatro primeiras páginas do documento.
    """
    if client.cnpj in only_digits(pdf_path.name):
        return True

    try:
        reader = PdfReader(str(pdf_path))

        for page in reader.pages[:4]:
            page_text = page.extract_text() or ""

            if client.cnpj in only_digits(page_text):
                return True
    except Exception as exc:
        LOGGER.warning(
            "Não foi possível validar o CNPJ dentro do PDF %s: %s",
            pdf_path.name,
            exc,
        )

    return False


def move_and_rename_report(
    downloaded_pdf: Path,
    client: Client,
) -> tuple[Path, str, bool]:
    """
    Cria a cópia final com nome confiável vindo da planilha.

    O original não é movido nem apagado. Isso mantém o item válido no
    Histórico do Chrome e evita o status "Excluído".
    """
    if not downloaded_pdf.exists():
        raise PortalError(
            f"O arquivo baixado não existe: {downloaded_pdf}"
        )

    if downloaded_pdf.stat().st_size <= 0:
        raise PortalError(
            f"O arquivo baixado está vazio: {downloaded_pdf.name}"
        )

    if not pdf_belongs_to_client(downloaded_pdf, client):
        raise PortalError(
            "O PDF baixado não parece pertencer ao CNPJ "
            f"{format_cnpj(client.cnpj)}. Arquivo recebido: "
            f"{downloaded_pdf.name}"
        )

    spreadsheet_name = client.spreadsheet_name.strip()
    used_fallback = not bool(spreadsheet_name)

    report_name = (
        spreadsheet_name
        or f"CNPJ {format_cnpj(client.cnpj)}"
    )

    clean_name = sanitize_filename(report_name)
    report_date = report_date_from_download_name(downloaded_pdf)

    final_filename = (
        sanitize_filename(
            (
                f"{clean_name} - "
                f"{format_cnpj(client.cnpj)} - "
                f"{report_date}"
            ),
            max_length=190,
        )
        + ".pdf"
    )

    final_path = OUTPUT_DIR / final_filename

    if final_path.exists():
        final_path.unlink()

    # Copia em vez de mover. O original permanece na pasta normal de downloads do Chrome.
    shutil.copy2(
        str(downloaded_pdf),
        str(final_path),
    )

    LOGGER.info(
        "Original preservado em: %s",
        downloaded_pdf,
    )
    LOGGER.info(
        "Cópia final criada em: %s",
        final_path,
    )

    return final_path, report_name, used_fallback


# =============================================================================
# PROCESSAMENTO DO LOOP
# =============================================================================

def represent_client(
    client: Client,
    learning_mode: bool,
) -> None:
    """
    Executa as etapas 1 a 8 do fluxo: abrir o painel, informar o CNPJ,
    selecionar Procurador, esperar 30 segundos, clicar em Representar e
    fechar o menu lateral.

    O bloco foi extraído de process_client sem qualquer alteração de
    comportamento para que outras automações da Central possam reaproveitar
    exatamente a mesma representação.
    """
    # Sempre começa no topo para deixar o fluxo previsível.
    keyboard_ctrl_home()

    # 1. Clica no perfil atual para abrir o painel.
    perform_mouse_action(
        "open_representation_panel",
        (
            "Clique no nome/perfil da empresa atualmente aberta, "
            "no canto superior direito, para abrir o painel Representar."
        ),
    )
    random_sleep(WAIT_AFTER_OPEN_PANEL_SECONDS)

    # 2. Clica no campo do CNPJ.
    perform_mouse_action(
        "cnpj_field",
        "Clique dentro do campo de CNPJ.",
    )

    # 3. Primeira empresa é digitada manualmente.
    #    As demais usam o ritmo aprendido.
    if learning_mode:
        capture_first_cnpj_and_typing_pattern(
            client.cnpj
        )
    else:
        type_cnpj_with_learned_pattern(
            client.cnpj
        )

    random_sleep(WAIT_AFTER_CNPJ_SECONDS)

    # 4. Primeiro clique para selecionar Procurador.
    perform_mouse_action(
        "profile_dropdown",
        (
            "Clique no campo de perfil onde deve ser "
            "selecionado 'Procurador'."
        ),
    )
    random_sleep(WAIT_AFTER_PROFILE_OPEN_SECONDS)

    # 5. Segundo clique para selecionar Procurador.
    perform_mouse_action(
        "procurador_option",
        "Clique na opção 'Procurador' da lista aberta.",
    )
    random_sleep(WAIT_AFTER_PROCURADOR_SECONDS)

    # 6. Antes de cada representação, espera obrigatoriamente 30 segundos.
    wait_before_represent()

    # 7. Representar.
    perform_mouse_action(
        "represent_button",
        "Agora clique no botão azul 'Representar'.",
    )

    LOGGER.info(
        "Aguardando 45 segundos após clicar em Representar para o portal carregar."
    )
    random_sleep(WAIT_AFTER_REPRESENT_SECONDS)

    # A nova página pode restaurar uma posição anterior. Volta ao topo antes
    # de procurar o X do menu lateral.
    keyboard_ctrl_home()

    # 8. Fecha o menu lateral após a representação ser carregada.
    # Isso devolve a página à largura normal e deixa o botão de download
    # sempre na mesma posição usada durante o aprendizado.
    perform_mouse_action(
        "close_representation_panel",
        (
            "Após a empresa representada terminar de carregar, "
            "clique no X do canto superior direito para fechar "
            "completamente o menu lateral."
        ),
    )
    random_sleep(WAIT_AFTER_CLOSE_PANEL_SECONDS)

    # O fechamento do menu pode alterar a altura/layout. Força novamente o topo.
    keyboard_ctrl_home()


def process_client(
    client: Client,
    learning_mode: bool,
) -> tuple[Path, str, bool]:
    if TEMP_DOWNLOAD_DIR is None:
        raise RuntimeError(
            "Diretório temporário não configurado."
        )

    represent_client(
        client,
        learning_mode=learning_mode,
    )

    # 9. Baixar o relatório.
    # Na segunda empresa, marca dez posições usando F8.
    # Da terceira em diante, testa essas dez posições até iniciar o download.
    if learning_mode:
        capture_all_download_positions()

    files_before = list_download_files(
        TEMP_DOWNLOAD_DIR
    )
    started_timestamp = time.time()

    click_download_positions_until_started(
        TEMP_DOWNLOAD_DIR,
        files_before,
        started_timestamp,
    )

    downloaded_pdf = wait_for_new_pdf(
        TEMP_DOWNLOAD_DIR,
        files_before,
        started_timestamp,
    )

    final_path, report_name, used_fallback = (
        move_and_rename_report(
            downloaded_pdf,
            client,
        )
    )

    LOGGER.info("PDF salvo: %s", final_path)

    return final_path, report_name, used_fallback




def process_first_client_fully_manual(
    client: Client,
) -> tuple[Path, str, bool]:
    """
    Processa somente a primeira empresa sem gravar cliques ou digitação.

    A execução é dividida em duas etapas para que o programa consiga impor
    a espera de 30 segundos antes de o usuário clicar em Representar.
    """
    if TEMP_DOWNLOAD_DIR is None:
        raise RuntimeError(
            "Diretório temporário não configurado."
        )

    # A primeira empresa é totalmente manual.
    # Nenhuma tecla de rolagem é enviada nesta fase.
    print()
    print("=" * 94)
    print("PRIMEIRA EMPRESA — FLUXO TOTALMENTE MANUAL")
    print("=" * 94)
    print(
        f"Empresa: {client.spreadsheet_name}"
    )
    print(
        f"CNPJ a informar: {format_cnpj(client.cnpj)}"
    )
    print()
    print("Faça manualmente somente estas etapas:")
    print("1. Clique no nome/perfil para abrir o menu lateral.")
    print("2. Clique no campo de CNPJ e informe o CNPJ acima.")
    print("3. Clique no campo de perfil.")
    print("4. Selecione 'Procurador'.")
    print()
    print(
        "NÃO clique em 'Representar' ainda. "
        "Depois de selecionar Procurador, volte ao terminal."
    )
    print("=" * 94)

    input(
        "\nPressione ENTER depois de preencher o CNPJ "
        "e selecionar Procurador..."
    )

    # A espera também é aplicada à primeira empresa, apesar de ela ser manual.
    wait_before_represent()

    files_before = list_download_files(
        TEMP_DOWNLOAD_DIR
    )
    started_timestamp = time.time()

    print()
    print("=" * 94)
    print("PRIMEIRA EMPRESA — CONTINUE MANUALMENTE")
    print("=" * 94)
    print("Agora faça estas etapas:")
    print("1. Clique em 'Representar'.")
    print("2. Aguarde a empresa terminar de carregar.")
    print("3. Clique no X para fechar o menu lateral.")
    print("4. Role a página, se necessário.")
    print("5. Clique em 'Baixar Relatório'.")
    print("6. Aguarde aparecer a confirmação de download.")
    print("7. Volte ao terminal e pressione ENTER.")
    print("=" * 94)

    input(
        "\nPressione ENTER somente depois de baixar o PDF "
        "da primeira empresa..."
    )

    if not wait_for_download_start(
        TEMP_DOWNLOAD_DIR,
        files_before,
        started_timestamp,
    ):
        raise PortalError(
            "O download manual da primeira empresa não foi encontrado."
        )

    downloaded_pdf = wait_for_new_pdf(
        TEMP_DOWNLOAD_DIR,
        files_before,
        started_timestamp,
    )

    final_path, report_name, used_fallback = (
        move_and_rename_report(
            downloaded_pdf,
            client,
        )
    )

    LOGGER.info(
        "Primeira empresa concluída manualmente. PDF salvo: %s",
        final_path,
    )

    return final_path, report_name, used_fallback


def learning_is_complete() -> bool:
    required_actions = set(REQUIRED_ACTION_KEYS)

    return (
        required_actions.issubset(MOUSE_ACTIONS)
        and len(TYPING_PATTERN) == 14
        and len(DOWNLOAD_POSITIONS) == DOWNLOAD_POSITION_COUNT
    )


# =============================================================================
# ENCERRAMENTO
# =============================================================================



# =============================================================================
# MAIN
# =============================================================================

def main() -> int:
    initialize_folders_and_logging()

    excel_path = resolve_excel_path()

    LOGGER.info("Planilha: %s", excel_path)
    LOGGER.info("Pasta de saída: %s", OUTPUT_DIR)

    clients = load_clients(excel_path)

    if not clients:
        LOGGER.warning(
            "Nenhum CNPJ ativo foi encontrado."
        )
        return 0

    LOGGER.info(
        "%d CNPJs serão processados. "
        "A META PLANO foi ignorada.",
        len(clients),
    )

    csv_handle, csv_writer = create_result_csv()

    successes = 0
    errors = 0
    fatal_error: Optional[str] = None

    try:
        if TEMP_DOWNLOAD_DIR is None:
            raise RuntimeError(
                "Diretório temporário não inicializado."
            )

        open_regular_chrome()

        LOGGER.info(
            "Pasta monitorada para novos PDFs: %s",
            TEMP_DOWNLOAD_DIR,
        )

        print()
        print("=" * 94)
        print("LOGIN MANUAL")
        print("=" * 94)
        print(
            "1. Use o Chrome REGULAR aberto pelo programa, com seu perfil normal."
        )
        print(
            "2. Faça o login pelo Gov.br usando o certificado da META PLANO."
        )
        print(
            "3. Aguarde o login ser concluído. O endereço interno será aberto "
            "somente depois que você pressionar ENTER."
        )
        print(
            "4. Deixe o Chrome maximizado no monitor que será usado."
        )
        print(
            "5. Volte ao terminal e pressione ENTER."
        )
        print()
        print(
            "A META PLANO não será processada. "
            "A primeira empresa será manual e a segunda será o aprendizado."
        )
        print("=" * 94)

        input(
            "\nPressione ENTER somente após concluir o login..."
        )

        LOGGER.info(
            "Login concluído. Abrindo 'Minhas Dívidas e Pendências' "
            "no Chrome já autenticado."
        )
        navigate_to_portal_with_keyboard()
        time.sleep(5)

        print()
        print("=" * 94)
        print("TRÊS FASES DO PROCESSAMENTO")
        print("=" * 94)
        print(
            "1ª empresa: você fará tudo manualmente; nada será gravado."
        )
        print(
            "2ª empresa: você fará o fluxo novamente e o programa "
            "gravará os cliques, o ritmo da digitação e dez posições "
            "possíveis do botão de download."
        )
        print(
            "3ª empresa em diante: o programa testará essas dez posições "
            "e parará assim que detectar o PDF."
        )
        print()
        print(
            "Em TODAS as empresas haverá uma espera obrigatória "
            "de 30 segundos antes de 'Representar'."
        )
        print("=" * 94)

        for position, client in enumerate(
            clients,
            start=1,
        ):
            started_at = datetime.now()

            # 1ª empresa: completamente manual, sem gravação.
            manual_first_mode = position == 1

            # 2ª empresa: grava todos os cliques e o ritmo da digitação.
            learning_mode = position == 2

            # 3ª empresa em diante: usa tudo o que foi aprendido.
            automatic_mode = position >= 3

            LOGGER.info(
                "[%d/%d] Linha %d | %s | %s",
                position,
                len(clients),
                client.excel_row,
                format_cnpj(client.cnpj),
                client.spreadsheet_name,
            )

            # A primeira empresa manual e a segunda empresa de aprendizado
            # não são repetidas automaticamente em caso de falha.
            max_attempts = (
                1
                if manual_first_mode or learning_mode
                else MAX_RETRIES_AFTER_LEARNING
            )

            completed = False
            last_message = ""

            for attempt in range(
                1,
                max_attempts + 1,
            ):
                try:
                    if manual_first_mode:
                        final_path, report_name, used_fallback = (
                            process_first_client_fully_manual(
                                client
                            )
                        )
                    else:
                        final_path, report_name, used_fallback = (
                            process_client(
                                client,
                                learning_mode=learning_mode,
                            )
                        )

                    if (
                        learning_mode
                        and not learning_is_complete()
                    ):
                        raise PortalError(
                            "O relatório da segunda empresa foi baixado, "
                            "mas os cliques, a digitação ou as cinco "
                            "posições de download não ficaram completos."
                        )

                    message = (
                        "Relatório baixado e renomeado com sucesso."
                    )

                    if used_fallback:
                        message += (
                            " Foi usado o nome da planilha porque "
                            "o PDF não forneceu uma razão social "
                            "reconhecível."
                        )

                    write_result(
                        csv_handle,
                        csv_writer,
                        client,
                        status="SUCESSO",
                        attempt=attempt,
                        message=message,
                        started_at=started_at,
                        report_name=report_name,
                        final_path=str(final_path),
                    )

                    successes += 1
                    completed = True

                    if manual_first_mode:
                        LOGGER.info(
                            "PRIMEIRA EMPRESA CONCLUÍDA MANUALMENTE. "
                            "A próxima empresa será usada para aprendizado."
                        )

                    if learning_mode:
                        LOGGER.info(
                            "APRENDIZADO DA SEGUNDA EMPRESA CONCLUÍDO. "
                            "O loop automático começará na terceira empresa."
                        )

                    break

                except Exception as exc:
                    last_message = str(exc)

                    LOGGER.error(
                        "Erro na linha %d, CNPJ %s, "
                        "tentativa %d/%d: %s",
                        client.excel_row,
                        format_cnpj(client.cnpj),
                        attempt,
                        max_attempts,
                        exc,
                    )

                    if manual_first_mode:
                        raise RuntimeError(
                            "A primeira empresa manual não foi concluída. "
                            "A execução foi interrompida antes do aprendizado."
                        ) from exc

                    if learning_mode:
                        raise RuntimeError(
                            "O aprendizado da segunda empresa não foi concluído. "
                            "A execução foi interrompida para não iniciar o loop "
                            "com posições incompletas."
                        ) from exc

                    if attempt < max_attempts:
                        LOGGER.info(
                            "Reabrindo a página do serviço antes "
                            "da nova tentativa."
                        )
                        navigate_to_portal_with_keyboard()

            if not completed:
                errors += 1

                write_result(
                    csv_handle,
                    csv_writer,
                    client,
                    status="ERRO",
                    attempt=max_attempts,
                    message=last_message or "Falha sem mensagem.",
                    started_at=started_at,
                )

                # Recupera a tela antes da empresa seguinte.
                LOGGER.info(
                    "Reabrindo a página do serviço antes "
                    "da próxima empresa."
                )
                navigate_to_portal_with_keyboard()

            if position < len(clients):
                delay = random.uniform(
                    *DELAY_BETWEEN_COMPANIES_SECONDS
                )

                LOGGER.info(
                    "Aguardando %.1f segundos antes "
                    "da próxima empresa.",
                    delay,
                )

                time.sleep(delay)

            if position % LONG_PAUSE_EVERY == 0:
                LOGGER.info(
                    "Pausa longa de %d segundos "
                    "após %d empresas.",
                    LONG_PAUSE_SECONDS,
                    position,
                )
                time.sleep(LONG_PAUSE_SECONDS)

    except KeyboardInterrupt:
        fatal_error = "Execução interrompida pelo usuário."
        LOGGER.warning(fatal_error)

    except Exception as exc:
        fatal_error = str(exc)
        LOGGER.critical(
            "Erro fatal: %s\n%s",
            exc,
            traceback.format_exc(),
        )

    finally:
        csv_handle.close()


    LOGGER.info(
        "Execução finalizada. "
        "Sucessos: %d | Erros: %d.",
        successes,
        errors,
    )

    if fatal_error:
        LOGGER.error(
            "Motivo da interrupção: %s",
            fatal_error,
        )
        return 1

    return 0 if errors == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
