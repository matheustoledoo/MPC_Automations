#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import importlib.util
import re
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

REQUIRED_PACKAGES = {
    "pypdf": "pypdf>=5.0",
    "openpyxl": "openpyxl>=3.1",
}


def ensure_dependencies() -> None:
    missing = [
        requirement
        for module_name, requirement in REQUIRED_PACKAGES.items()
        if importlib.util.find_spec(module_name) is None
    ]
    if missing:
        print("[SETUP] Instalando dependências: " + ", ".join(missing), flush=True)
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", *missing])


ensure_dependencies()

from openpyxl import Workbook  # noqa: E402
from openpyxl.chart import BarChart, Reference  # noqa: E402
from openpyxl.formatting.rule import CellIsRule, FormulaRule  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402
from pypdf import PdfReader  # noqa: E402


MAROON = "8B0000"
MAROON_DARK = "5F0000"
GOLD = "D4AF37"
GRAY_DARK = "404040"
GRAY = "666666"
GRAY_LIGHT = "F2F2F2"
WHITE = "FFFFFF"
RED_LIGHT = "FCE8E6"
GREEN_LIGHT = "EAF4EA"
YELLOW_LIGHT = "FFF4CC"
BLUE_LIGHT = "EAF2F8"

MONTHS = ("JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ")
MONEY_RE = re.compile(r"(?<!\d)-?\d{1,3}(?:\.\d{3})*,\d{2}|(?<!\d)-?\d+,\d{2}")
SIDA_ID_RE = re.compile(r"\d{2}\.\d\.\d{2}\.\d{6}-\d{2}")
PROCESS_RE = re.compile(r"\d{5}\.\d{3}\.\d{3}/\d{4}-\d{2}")
DATE_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")

BASE_HEADERS = [
    "Razão Social",
    "CNPJ",
    "Situação",
    "Situação Completa",
    "Data/Hora do Relatório",
    "UA de Domicílio",
    "Código da UA",
    "Endereço",
    "Bairro",
    "CEP",
    "Município",
    "UF",
    "Responsável",
    "CPF do Responsável",
    "Natureza Jurídica",
    "Data de Abertura",
    "CNAE",
    "Porte da Empresa",
    "Há seção 'Opção pelo Simples Nacional'",
]

FISCAL_HEADERS = [
    "Qtd. Tipos de Omissão",
    "Qtd. Períodos Omitidos",
    "Tipos de Omissão",
    "Detalhamento das Omissões",
    "Qtd. Tipos de Pendência",
    "Qtd. Itens de Pendência",
    "Tipos de Pendência",
    "Detalhamento das Pendências",
    "Valor Original Total (R$)",
    "Saldo Devedor Total (R$)",
    "Multa Total (R$)",
    "Juros Total (R$)",
    "Saldo Consolidado Total (R$)",
    "Qtd. Inscrições SIDA",
]

FILE_HEADERS = ["Arquivo PDF", "Caminho do PDF"]
MAIN_HEADERS = BASE_HEADERS + FISCAL_HEADERS + FILE_HEADERS

OMISSION_DETAIL_HEADERS = [
    "Razão Social",
    "CNPJ",
    "Situação",
    "Origem",
    "Tipo de Omissão",
    "Períodos/Ano-calendário",
    "Quantidade de Períodos",
    "Detalhamento Original",
    "Arquivo PDF",
    "Caminho do PDF",
]

PENDENCY_DETAIL_HEADERS = [
    "Razão Social",
    "CNPJ",
    "Situação",
    "Origem",
    "Tipo de Pendência",
    "Descrição/Referência",
    "Inscrição SIDA",
    "Processo",
    "Situação do Item",
    "Valor Original (R$)",
    "Saldo Devedor (R$)",
    "Multa (R$)",
    "Juros (R$)",
    "Saldo Consolidado (R$)",
    "Detalhamento Original",
    "Arquivo PDF",
    "Caminho do PDF",
]


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", value).strip().casefold()


def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def brl_to_float(value: str) -> float:
    if not value:
        return 0.0
    cleaned = value.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def format_brl(value: float) -> str:
    formatted = f"{value:,.2f}"
    return "R$ " + formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def read_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    parts: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text:
            parts.append(text)
    return "\n".join(parts)


def find_line_value(text: str, label_regex: str) -> str:
    pattern = rf"(?im)^\s*{label_regex}\s*:\s*(.*?)\s*$"
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def detect_cadastral_status(text: str) -> tuple[str, str]:
    """Retorna a classificação curta e o texto completo do campo Situação."""
    match = re.search(
        r"(?im)^\s*Situa[cç][aã]o(?:\s+Cadastral)?\s*:\s*([^\r\n]+)",
        text,
    )
    if not match:
        return "", ""

    raw = re.sub(r"\s+", " ", match.group(1)).strip(" _")
    normalized = normalize_text(raw)

    if re.match(r"^inapt[ao]\b", normalized):
        return "INAPTA", raw
    if re.match(r"^ativ[ao]\b", normalized):
        return "ATIVA", raw
    return raw.split()[0].upper() if raw else "", raw


def extract_company_data(pdf_path: Path, text: str, status: str, status_raw: str) -> dict[str, Any]:
    company = ""
    match = re.search(
        r"(?im)^\s*CNPJ:\s*\d{2}\.\d{3}\.\d{3}\s*-\s*(.*?)\s*$",
        text,
    )
    if match:
        company = match.group(1).strip()

    cnpj = ""
    cadastral_match = re.search(
        r"(?is)Dados\s+Cadastrais\s+da\s+Matriz.*?CNPJ:\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})",
        text,
    )
    if cadastral_match:
        cnpj = cadastral_match.group(1)
    else:
        all_cnpjs = re.findall(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", text)
        if all_cnpjs:
            cnpj = all_cnpjs[-1]

    report_datetime = ""
    match = re.search(r"\b(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})\b", text)
    if match:
        report_datetime = match.group(1)

    ua = ""
    ua_code = ""
    match = re.search(
        r"(?im)^\s*UA\s+de\s+Domic[ií]lio:\s*(.*?)\s+C[oó]digo\s+da\s+UA:\s*(.*?)\s*$",
        text,
    )
    if match:
        ua = match.group(1).strip()
        ua_code = match.group(2).strip()

    address = find_line_value(text, r"Endere[cç]o")

    bairro = cep = municipio = uf = ""
    match = re.search(
        r"(?i)Bairro:\s*([^\r\n]*?)\s+CEP:\s*([\d.\-]+)\s*Munic[ií]pio:\s*([^\r\n]*?)\s+UF:\s*([A-Z]{2})",
        text,
    )
    if match:
        bairro, cep, municipio, uf = [item.strip() for item in match.groups()]

    responsible_cpf = responsible_name = ""
    match = re.search(
        r"(?im)^\s*Respons[aá]vel:\s*([\d.\-]+)\s*-\s*(.*?)\s*$",
        text,
    )
    if match:
        responsible_cpf = match.group(1).strip()
        responsible_name = match.group(2).strip()

    legal_nature = opening_date = ""
    match = re.search(
        r"(?im)^\s*Natureza\s+Jur[ií]dica:\s*(.*?)\s+Data\s+de\s+Abertura:\s*(\d{2}/\d{2}/\d{4})\s*$",
        text,
    )
    if match:
        legal_nature = match.group(1).strip()
        opening_date = match.group(2).strip()

    return {
        "Razão Social": company,
        "CNPJ": cnpj,
        "Situação": status,
        "Situação Completa": status_raw,
        "Data/Hora do Relatório": report_datetime,
        "UA de Domicílio": ua,
        "Código da UA": ua_code,
        "Endereço": address,
        "Bairro": bairro,
        "CEP": cep,
        "Município": municipio,
        "UF": uf,
        "Responsável": responsible_name,
        "CPF do Responsável": responsible_cpf,
        "Natureza Jurídica": legal_nature,
        "Data de Abertura": opening_date,
        "CNAE": find_line_value(text, r"CNAE"),
        "Porte da Empresa": find_line_value(text, r"Porte\s+da\s+Empresa"),
        "Há seção 'Opção pelo Simples Nacional'": (
            "SIM" if re.search(r"(?i)Op[cç][aã]o\s+pelo\s+Simples\s+Nacional", text) else ""
        ),
        "Arquivo PDF": pdf_path.name,
        "Caminho do PDF": str(pdf_path.resolve()),
    }


def clean_pdf_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw in text.splitlines():
        line = re.sub(r"\s+", " ", raw).strip()
        if not line:
            continue
        normalized = normalize_text(line)

        if normalized.startswith("ministerio da fazenda"):
            continue
        if normalized.startswith("secretaria especial da receita federal"):
            continue
        if normalized.startswith("procuradoria-geral da fazenda nacional"):
            continue
        if normalized.startswith("informacoes de apoio para emissao"):
            continue
        if normalized.startswith("pagina:") or re.fullmatch(r"\d+", normalized):
            continue
        if normalized.startswith("cnpj do certificado:"):
            continue
        if re.match(r"^cnpj:\s*\d{2}\.\d{3}\.\d{3}\s*-", line, flags=re.IGNORECASE):
            continue
        if "final do relatorio" in normalized:
            continue

        lines.append(line)
    return lines


def clean_heading(line: str) -> str:
    return re.sub(r"_+", "", line).strip(" -_")


def extract_fiscal_sections(text: str) -> list[dict[str, Any]]:
    lines = clean_pdf_lines(text)
    sections: list[dict[str, Any]] = []
    current: Optional[dict[str, Any]] = None
    origin = "Receita Federal"

    def finish_current() -> None:
        nonlocal current
        if current is not None:
            current["details"] = [item for item in current["details"] if item]
            sections.append(current)
            current = None

    for line in lines:
        normalized = normalize_text(line)

        if "diagnostico fiscal na procuradoria-geral" in normalized:
            finish_current()
            origin = "PGFN"
            continue
        if "diagnostico fiscal na receita federal" in normalized:
            finish_current()
            origin = "Receita Federal"
            continue

        heading = clean_heading(line)
        heading_normalized = normalize_text(heading)

        if re.match(r"^omissao\b", heading_normalized):
            finish_current()
            current = {
                "kind": "OMISSÃO",
                "title": heading,
                "origin": origin,
                "details": [],
            }
            continue

        if re.match(r"^pendencia\b", heading_normalized):
            finish_current()
            current = {
                "kind": "PENDÊNCIA",
                "title": heading,
                "origin": origin,
                "details": [],
            }
            continue

        if current is not None:
            # CNPJ e títulos repetidos de tabelas não agregam informação ao detalhe.
            if re.fullmatch(r"CNPJ:\s*\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", line, flags=re.IGNORECASE):
                continue
            current["details"].append(line)

    finish_current()
    return sections


def omission_type(title: str) -> str:
    cleaned = re.sub(
        r"(?i)^Omiss[aã]o\s*(?:de|da|do)?\s*",
        "",
        clean_heading(title),
    ).strip()
    return cleaned or clean_heading(title)


def pendency_type(title: str) -> str:
    cleaned = re.sub(
        r"(?i)^Pend[eê]ncia\s*-?\s*",
        "",
        clean_heading(title),
    ).strip()
    return cleaned or clean_heading(title)


def omission_period_count(details: list[str]) -> int:
    total = 0
    found_period_information = False

    for line in details:
        normalized_upper = normalize_text(line).upper()
        months = re.findall(r"\b(?:" + "|".join(MONTHS) + r")\b", normalized_upper)
        if months:
            found_period_information = True
            total += len(months)
            continue

        if "ANO-CALENDARIO" in normalized_upper or "PERIODO DE APURACAO" in normalized_upper:
            years = re.findall(r"\b20\d{2}\b", line)
            if years:
                found_period_information = True
                total += len(set(years))

    if found_period_information and total > 0:
        return total
    return 1


def parse_omission_section(
    section: dict[str, Any],
    company: dict[str, Any],
) -> dict[str, Any]:
    details = section.get("details", [])
    detail_text = " | ".join(details)
    count = omission_period_count(details)

    return {
        "Razão Social": company["Razão Social"],
        "CNPJ": company["CNPJ"],
        "Situação": company["Situação"],
        "Origem": section.get("origin", ""),
        "Tipo de Omissão": omission_type(section.get("title", "")),
        "Períodos/Ano-calendário": detail_text,
        "Quantidade de Períodos": count,
        "Detalhamento Original": detail_text,
        "Arquivo PDF": company["Arquivo PDF"],
        "Caminho do PDF": company["Caminho do PDF"],
    }


def parse_sief_rows(
    details: list[str],
    section: dict[str, Any],
    company: dict[str, Any],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for index, line in enumerate(details):
        normalized = normalize_text(line)
        if normalized.startswith("receita pa/exerc"):
            continue

        money_matches = list(MONEY_RE.finditer(line))
        if len(money_matches) < 5:
            continue

        selected = money_matches[-5:]
        amounts = [brl_to_float(match.group(0)) for match in selected]
        description = line[: selected[0].start()].strip(" -")
        item_status = line[selected[-1].end() :].strip(" -")
        raw_detail = line

        if index + 1 < len(details) and normalize_text(details[index + 1]).startswith("notificacao de lancamento"):
            raw_detail += " | " + details[index + 1]

        rows.append({
            "Razão Social": company["Razão Social"],
            "CNPJ": company["CNPJ"],
            "Situação": company["Situação"],
            "Origem": section.get("origin", ""),
            "Tipo de Pendência": pendency_type(section.get("title", "")),
            "Descrição/Referência": description,
            "Inscrição SIDA": "",
            "Processo": "",
            "Situação do Item": item_status,
            "Valor Original (R$)": amounts[0],
            "Saldo Devedor (R$)": amounts[1],
            "Multa (R$)": amounts[2],
            "Juros (R$)": amounts[3],
            "Saldo Consolidado (R$)": amounts[4],
            "Detalhamento Original": raw_detail,
            "Arquivo PDF": company["Arquivo PDF"],
            "Caminho do PDF": company["Caminho do PDF"],
        })

    return rows


def parse_sida_rows(
    details: list[str],
    section: dict[str, Any],
    company: dict[str, Any],
) -> list[dict[str, Any]]:
    blocks: list[list[str]] = []
    current: list[str] = []

    for line in details:
        if SIDA_ID_RE.search(line):
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)

    if current:
        blocks.append(current)

    rows: list[dict[str, Any]] = []
    for block in blocks:
        raw_detail = " | ".join(block)
        inscription_match = SIDA_ID_RE.search(raw_detail)
        process_match = PROCESS_RE.search(raw_detail)
        status_match = re.search(r"(?i)Situa[cç][aã]o:\s*(.*?)(?:\s*_{2,}|$)", raw_detail)

        description = raw_detail
        if inscription_match:
            description = raw_detail.replace(inscription_match.group(0), "", 1).strip(" |-")

        rows.append({
            "Razão Social": company["Razão Social"],
            "CNPJ": company["CNPJ"],
            "Situação": company["Situação"],
            "Origem": section.get("origin", ""),
            "Tipo de Pendência": pendency_type(section.get("title", "")),
            "Descrição/Referência": description,
            "Inscrição SIDA": inscription_match.group(0) if inscription_match else "",
            "Processo": process_match.group(0) if process_match else "",
            "Situação do Item": status_match.group(1).strip() if status_match else "",
            "Valor Original (R$)": 0.0,
            "Saldo Devedor (R$)": 0.0,
            "Multa (R$)": 0.0,
            "Juros (R$)": 0.0,
            "Saldo Consolidado (R$)": 0.0,
            "Detalhamento Original": raw_detail,
            "Arquivo PDF": company["Arquivo PDF"],
            "Caminho do PDF": company["Caminho do PDF"],
        })

    return rows


def parse_generic_pendency(
    details: list[str],
    section: dict[str, Any],
    company: dict[str, Any],
) -> list[dict[str, Any]]:
    detail_text = " | ".join(details).strip()
    return [{
        "Razão Social": company["Razão Social"],
        "CNPJ": company["CNPJ"],
        "Situação": company["Situação"],
        "Origem": section.get("origin", ""),
        "Tipo de Pendência": pendency_type(section.get("title", "")),
        "Descrição/Referência": detail_text,
        "Inscrição SIDA": "",
        "Processo": "",
        "Situação do Item": "",
        "Valor Original (R$)": 0.0,
        "Saldo Devedor (R$)": 0.0,
        "Multa (R$)": 0.0,
        "Juros (R$)": 0.0,
        "Saldo Consolidado (R$)": 0.0,
        "Detalhamento Original": detail_text,
        "Arquivo PDF": company["Arquivo PDF"],
        "Caminho do PDF": company["Caminho do PDF"],
    }]


def parse_pendency_section(
    section: dict[str, Any],
    company: dict[str, Any],
) -> list[dict[str, Any]]:
    title_normalized = normalize_text(section.get("title", ""))
    details = section.get("details", [])

    sief_rows = parse_sief_rows(details, section, company)
    if sief_rows:
        return sief_rows

    if "sida" in title_normalized or any(SIDA_ID_RE.search(line) for line in details):
        sida_rows = parse_sida_rows(details, section, company)
        if sida_rows:
            return sida_rows

    return parse_generic_pendency(details, section, company)


def summarize_fiscal_data(
    company: dict[str, Any],
    omissions: list[dict[str, Any]],
    pendencies: list[dict[str, Any]],
) -> dict[str, Any]:
    omission_types = list(dict.fromkeys(item["Tipo de Omissão"] for item in omissions))
    pendency_types = list(dict.fromkeys(item["Tipo de Pendência"] for item in pendencies))

    omission_summary = "\n".join(
        f"• {item['Tipo de Omissão']}: {item['Períodos/Ano-calendário']} "
        f"({item['Quantidade de Períodos']} período(s))"
        for item in omissions
    )

    pendency_lines: list[str] = []
    for pendency_name in pendency_types:
        matching = [item for item in pendencies if item["Tipo de Pendência"] == pendency_name]
        consolidated = sum(float(item["Saldo Consolidado (R$)"] or 0) for item in matching)
        sida_count = sum(1 for item in matching if item["Inscrição SIDA"])
        extra = ""
        if consolidated:
            extra = f"; saldo consolidado {format_brl(consolidated)}"
        elif sida_count:
            extra = f"; {sida_count} inscrição(ões) SIDA"
        pendency_lines.append(f"• {pendency_name}: {len(matching)} item(ns){extra}")

    fiscal = {
        "Qtd. Tipos de Omissão": len(omission_types),
        "Qtd. Períodos Omitidos": sum(int(item["Quantidade de Períodos"] or 0) for item in omissions),
        "Tipos de Omissão": " | ".join(omission_types),
        "Detalhamento das Omissões": omission_summary,
        "Qtd. Tipos de Pendência": len(pendency_types),
        "Qtd. Itens de Pendência": len(pendencies),
        "Tipos de Pendência": " | ".join(pendency_types),
        "Detalhamento das Pendências": "\n".join(pendency_lines),
        "Valor Original Total (R$)": sum(float(item["Valor Original (R$)"] or 0) for item in pendencies),
        "Saldo Devedor Total (R$)": sum(float(item["Saldo Devedor (R$)"] or 0) for item in pendencies),
        "Multa Total (R$)": sum(float(item["Multa (R$)"] or 0) for item in pendencies),
        "Juros Total (R$)": sum(float(item["Juros (R$)"] or 0) for item in pendencies),
        "Saldo Consolidado Total (R$)": sum(float(item["Saldo Consolidado (R$)"] or 0) for item in pendencies),
        "Qtd. Inscrições SIDA": sum(1 for item in pendencies if item["Inscrição SIDA"]),
    }

    return {**company, **fiscal}


def analyze_folder(
    pdf_dir: Path,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    int,
    list[str],
]:
    pdfs = sorted(pdf_dir.glob("*.pdf"), key=lambda path: path.name.casefold())
    inapt_rows: list[dict[str, Any]] = []
    active_issue_rows: list[dict[str, Any]] = []
    omission_details: list[dict[str, Any]] = []
    pendency_details: list[dict[str, Any]] = []
    errors: list[str] = []

    total = len(pdfs)
    print(f"[ANÁLISE] {total} PDF(s) encontrado(s) em: {pdf_dir}", flush=True)

    for index, pdf_path in enumerate(pdfs, start=1):
        print(f"[SCAN {index}/{total}] {pdf_path.name}", flush=True)
        try:
            text = read_pdf_text(pdf_path)
            if not text.strip():
                errors.append(f"{pdf_path.name}: PDF sem texto extraível")
                print(f"[AVISO] PDF sem texto extraível: {pdf_path.name}", flush=True)
                continue

            status, status_raw = detect_cadastral_status(text)
            if not status:
                errors.append(f"{pdf_path.name}: campo Situação cadastral não localizado")
                print(f"[AVISO] Situação cadastral não localizada: {pdf_path.name}", flush=True)
                continue

            company = extract_company_data(pdf_path, text, status, status_raw)
            sections = extract_fiscal_sections(text)

            company_omissions: list[dict[str, Any]] = []
            company_pendencies: list[dict[str, Any]] = []

            for section in sections:
                if section["kind"] == "OMISSÃO":
                    company_omissions.append(parse_omission_section(section, company))
                elif section["kind"] == "PENDÊNCIA":
                    company_pendencies.extend(parse_pendency_section(section, company))

            row = summarize_fiscal_data(company, company_omissions, company_pendencies)
            omission_details.extend(company_omissions)
            pendency_details.extend(company_pendencies)

            has_fiscal_issue = bool(company_omissions or company_pendencies)
            if status == "INAPTA":
                inapt_rows.append(row)
                print(
                    f"[INAPTA] {company.get('Razão Social') or pdf_path.stem} | "
                    f"{company.get('CNPJ')} | "
                    f"{len(company_omissions)} omissão(ões) | "
                    f"{len(company_pendencies)} pendência(s)",
                    flush=True,
                )
            elif status == "ATIVA" and has_fiscal_issue:
                active_issue_rows.append(row)
                print(
                    f"[ATIVA COM PENDÊNCIAS] {company.get('Razão Social') or pdf_path.stem} | "
                    f"{company.get('CNPJ')} | "
                    f"{len(company_omissions)} omissão(ões) | "
                    f"{len(company_pendencies)} pendência(s)",
                    flush=True,
                )

        except Exception as exc:
            errors.append(f"{pdf_path.name}: {exc}")
            print(f"[ERRO] {pdf_path.name}: {exc}", flush=True)

    return inapt_rows, active_issue_rows, omission_details, pendency_details, total, errors


def style_title(ws, title: str, subtitle: str, total_columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=MAROON)
    cell.font = Font(color=WHITE, bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    sub = ws.cell(2, 1, subtitle)
    sub.font = Font(color=GRAY, italic=True, size=10)
    sub.alignment = Alignment(horizontal="left", vertical="center")


def style_headers(ws, headers: list[str], header_row: int) -> None:
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, column, header)
        cell.fill = PatternFill("solid", fgColor=GRAY_DARK)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[header_row].height = 44


def add_hyperlinks(ws, headers: list[str], first_data_row: int, last_row: int) -> None:
    if "Caminho do PDF" not in headers:
        return
    column = headers.index("Caminho do PDF") + 1
    for row in range(first_data_row, last_row + 1):
        cell = ws.cell(row, column)
        if not cell.value:
            continue
        try:
            cell.hyperlink = Path(str(cell.value)).as_uri()
            cell.style = "Hyperlink"
        except Exception:
            pass


def add_table(ws, headers: list[str], header_row: int, row_count: int, table_name: str) -> None:
    if row_count <= 0:
        return
    last_row = header_row + row_count
    last_column = get_column_letter(len(headers))
    table = Table(displayName=table_name, ref=f"A{header_row}:{last_column}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[dict[str, Any]],
    table_name: str,
    status_fill: str,
    widths: Optional[dict[str, float]] = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    style_title(ws, title, subtitle, len(headers))
    header_row = 4
    style_headers(ws, headers, header_row)

    for row_index, item in enumerate(rows, start=header_row + 1):
        for column_index, header in enumerate(headers, start=1):
            value = item.get(header, "")
            cell = ws.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")

            if header.endswith("(R$)"):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif header.startswith("Qtd.") or header == "Quantidade de Períodos":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="top")

        if "Situação" in headers:
            status_column = headers.index("Situação") + 1
            status_cell = ws.cell(row_index, status_column)
            status_cell.fill = PatternFill("solid", fgColor=status_fill)
            status_cell.font = Font(color=MAROON_DARK, bold=True)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = header_row + len(rows)
    add_table(ws, headers, header_row, len(rows), table_name)
    add_hyperlinks(ws, headers, header_row + 1, last_row)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row + 1, last_row)}"

    if widths:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
    else:
        for column_index, header in enumerate(headers, start=1):
            letter = get_column_letter(column_index)
            if "Detalhamento" in header or "Caminho" in header:
                ws.column_dimensions[letter].width = 52
            elif "Razão Social" in header or "Descrição" in header:
                ws.column_dimensions[letter].width = 38
            elif "Tipo" in header or "Natureza" in header or "CNAE" in header:
                ws.column_dimensions[letter].width = 30
            elif "Endereço" in header:
                ws.column_dimensions[letter].width = 32
            elif header.endswith("(R$)"):
                ws.column_dimensions[letter].width = 18
            else:
                ws.column_dimensions[letter].width = 18

    if rows and "Saldo Consolidado Total (R$)" in headers:
        column = headers.index("Saldo Consolidado Total (R$)") + 1
        letter = get_column_letter(column)
        ws.conditional_formatting.add(
            f"{letter}{header_row + 1}:{letter}{last_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor=YELLOW_LIGHT),
            ),
        )


def build_excel(
    output_path: Path,
    inapt_rows: list[dict[str, Any]],
    active_issue_rows: list[dict[str, Any]],
    omission_details: list[dict[str, Any]],
    pendency_details: list[dict[str, Any]],
    total_scanned: int,
    errors: list[str],
    source_dir: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary = wb.active
    summary.title = "Resumo"
    summary.sheet_view.showGridLines = False

    summary.merge_cells("A1:H2")
    title = summary["A1"]
    title.value = "META PLANO | DIAGNÓSTICO FISCAL DOS RELATÓRIOS"
    title.fill = PatternFill("solid", fgColor=MAROON)
    title.font = Font(color=WHITE, bold=True, size=18)
    title.alignment = Alignment(horizontal="left", vertical="center")

    metrics = [
        ("PDFs analisados", total_scanned),
        ("Empresas INAPTAS", len(inapt_rows)),
        ("ATIVAS com pendências/omissões", len(active_issue_rows)),
        ("Omissões encontradas", len(omission_details)),
        ("Itens de pendência", len(pendency_details)),
        ("Erros de leitura", len(errors)),
    ]

    positions = [(4, 1), (4, 3), (4, 5), (7, 1), (7, 3), (7, 5)]
    for (label, value), (row, column) in zip(metrics, positions):
        summary.cell(row, column, label)
        summary.cell(row, column).font = Font(bold=True, color=GRAY_DARK, size=10)
        summary.cell(row + 1, column, value)
        summary.cell(row + 1, column).font = Font(bold=True, color=MAROON, size=22)
        summary.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        summary.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)

    total_consolidated = sum(float(item["Saldo Consolidado (R$)"] or 0) for item in pendency_details)
    summary["A10"] = "Saldo consolidado identificado"
    summary["A10"].font = Font(bold=True, color=GRAY_DARK)
    summary["A11"] = total_consolidated
    summary["A11"].number_format = 'R$ #,##0.00'
    summary["A11"].font = Font(bold=True, color=MAROON, size=20)

    summary["A13"] = "Pasta analisada"
    summary["B13"] = str(source_dir.resolve())
    summary["A14"] = "Relatório gerado em"
    summary["B14"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    summary["A16"] = "Critérios"
    summary["B16"] = (
        "Empresas INAPTAS são sempre incluídas. Empresas ATIVAS entram na aba específica "
        "quando o Diagnóstico Fiscal contém ao menos uma seção iniciada por Omissão ou Pendência. "
        "Os nomes posteriores podem variar; a classificação usa somente essas palavras-chave."
    )
    summary["B16"].alignment = Alignment(wrap_text=True, vertical="top")

    if errors:
        summary["A19"] = "Avisos de leitura"
        summary["A19"].font = Font(bold=True, color=MAROON)
        for row_index, error in enumerate(errors[:30], start=20):
            summary.cell(row_index, 1, f"• {error}")
            summary.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
            summary.cell(row_index, 1).alignment = Alignment(wrap_text=True)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 42
    summary.column_dimensions["C"].width = 26
    summary.column_dimensions["D"].width = 8
    summary.column_dimensions["E"].width = 24
    summary.column_dimensions["F"].width = 16
    summary.column_dimensions["G"].width = 16
    summary.column_dimensions["H"].width = 16

    # Pequeno gráfico visual dos principais grupos.
    summary["J3"] = "Categoria"
    summary["K3"] = "Quantidade"
    chart_data = [
        ("INAPTAS", len(inapt_rows)),
        ("ATIVAS com pendências", len(active_issue_rows)),
        ("Omissões", len(omission_details)),
        ("Pendências", len(pendency_details)),
    ]
    for index, (label, value) in enumerate(chart_data, start=4):
        summary.cell(index, 10, label)
        summary.cell(index, 11, value)
    chart = BarChart()
    chart.title = "Resumo do diagnóstico"
    chart.y_axis.title = "Quantidade"
    chart.height = 7
    chart.width = 12
    chart.add_data(Reference(summary, min_col=11, min_row=3, max_row=7), titles_from_data=True)
    chart.set_categories(Reference(summary, min_col=10, min_row=4, max_row=7))
    chart.legend = None
    summary.add_chart(chart, "J9")

    write_data_sheet(
        wb,
        "Empresas Inaptas",
        "EMPRESAS COM SITUAÇÃO CADASTRAL INAPTA",
        f"{len(inapt_rows)} empresa(s) INAPTA(S) encontrada(s) em {total_scanned} PDF(s).",
        MAIN_HEADERS,
        inapt_rows,
        "EmpresasInaptasTable",
        RED_LIGHT,
    )

    write_data_sheet(
        wb,
        "Ativas com Pendências",
        "EMPRESAS ATIVAS COM OMISSÕES E/OU PENDÊNCIAS",
        f"{len(active_issue_rows)} empresa(s) ATIVA(S) com diagnóstico fiscal relevante.",
        MAIN_HEADERS,
        active_issue_rows,
        "AtivasPendenciasTable",
        GREEN_LIGHT,
    )

    write_data_sheet(
        wb,
        "Omissões Detalhadas",
        "OMISSÕES DETALHADAS",
        "Uma linha por tipo de omissão. A quantidade considera os meses ou anos informados no relatório.",
        OMISSION_DETAIL_HEADERS,
        omission_details,
        "OmissoesDetalhadasTable",
        YELLOW_LIGHT,
    )

    write_data_sheet(
        wb,
        "Pendências Detalhadas",
        "PENDÊNCIAS DETALHADAS",
        "Uma linha por débito, inscrição SIDA ou outra ocorrência identificada no diagnóstico fiscal.",
        PENDENCY_DETAIL_HEADERS,
        pendency_details,
        "PendenciasDetalhadasTable",
        YELLOW_LIGHT,
    )

    wb.save(output_path)
    print(f"[EXCEL] Arquivo criado: {output_path}", flush=True)
    print(f"[RESUMO] PDFs analisados: {total_scanned}", flush=True)
    print(f"[RESUMO] Empresas INAPTAS: {len(inapt_rows)}", flush=True)
    print(f"[RESUMO] Empresas ATIVAS com pendências/omissões: {len(active_issue_rows)}", flush=True)
    print(f"[RESUMO] Omissões detalhadas: {len(omission_details)}", flush=True)
    print(f"[RESUMO] Itens de pendência: {len(pendency_details)}", flush=True)
    if errors:
        print(f"[RESUMO] PDFs com aviso/erro de leitura: {len(errors)}", flush=True)


def run(pdf_dir: Path, output_path: Path) -> int:
    if not pdf_dir.exists() or not pdf_dir.is_dir():
        print(f"[ERRO] Pasta de PDFs inválida: {pdf_dir}", flush=True)
        return 2

    (
        inapt_rows,
        active_issue_rows,
        omission_details,
        pendency_details,
        total,
        errors,
    ) = analyze_folder(pdf_dir)

    build_excel(
        output_path,
        inapt_rows,
        active_issue_rows,
        omission_details,
        pendency_details,
        total,
        errors,
        pdf_dir,
    )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Analisa situação cadastral, omissões e pendências nos PDFs da Receita Federal."
        )
    )
    parser.add_argument("--pdf-dir", required=True, help="Pasta contendo somente os PDFs finais.")
    parser.add_argument("--output", required=True, help="Arquivo .xlsx de saída.")
    args = parser.parse_args()

    return run(Path(args.pdf_dir), Path(args.output))


if __name__ == "__main__":
    raise SystemExit(main())
